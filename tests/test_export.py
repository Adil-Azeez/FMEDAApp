import unittest
import tempfile
import os
import openpyxl
from fmeda_tool.models import Project, Unit, Component, FailureModeAssignment
from fmeda_tool.services import ExportService, ProjectService


class TestExportService(unittest.TestCase):
    
    def setUp(self):
        # Create a mock project
        self.project = Project(
            id="proj_export_01",
            name="Test Export Project",
            description="Export Desc",
            project_number="EXP-001",
            version="1.2.3",
            target_sil="SIL 2",
            reviewer="Sarah Connor",
            custom_fields={"Subsystem ID": "SUB-99"}
        )
        
        # Add a unit and component
        self.unit = Unit(
            id="unit_01",
            name="Power Supply",
            description="Power regulation board"
        )
        
        self.comp = Component(
            id="comp_01",
            position="R101",
            name="Resistor",
            type="Resistor",
            failure_rate=0.5,
            failure_modes={"Short": 60.0, "Open": 40.0}
        )
        
        self.assignment1 = FailureModeAssignment(
            failure_mode_name="Short",
            failure_rate_percentage=60.0,
            classification="dangerous_failure",
            deviation_id="dev_001",
            notes="Gate shorted"
        )
        self.assignment2 = FailureModeAssignment(
            failure_mode_name="Open",
            failure_rate_percentage=40.0,
            classification="safe_failure",
            notes="Gate open circuit"
        )
        self.comp.failure_mode_assignments = [self.assignment1, self.assignment2]
        self.unit.components = [self.comp]
        self.project.units = [self.unit]
        
        # Log a change history entry
        ProjectService.log_change(self.project, "Create Component", "Added R101", "Engineer Bob")
        
    def test_export_excel_with_all_options(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            file_path = os.path.join(tmpdir, "export_test.xlsx")
            
            # Export
            success = ExportService.export_to_excel(
                self.project,
                file_path,
                include_summary=True,
                include_custom=True,
                include_history=True,
                fmeda_per_fg=True
            )
            self.assertTrue(success)
            self.assertTrue(os.path.exists(file_path))
            
            # Verify worksheets in saved workbook
            wb = openpyxl.load_workbook(file_path)
            self.assertIn("Overview", wb.sheetnames)
            self.assertIn("Power Supply", wb.sheetnames)
            self.assertIn("Change History", wb.sheetnames)
            
            # Verify column count in FMEDA sheet (39 columns)
            ws_fg = wb["Power Supply"]
            self.assertEqual(ws_fg.max_column, 39)
            self.assertEqual(ws_fg.cell(row=1, column=1).value, "Component ID / Designator")
            
            # Verify values
            self.assertEqual(ws_fg.cell(row=2, column=1).value, "R101")
            self.assertEqual(ws_fg.cell(row=2, column=8).value, "Short")
            self.assertEqual(ws_fg.cell(row=3, column=8).value, "Open")
            
    def test_export_excel_combined_fmeda(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            file_path = os.path.join(tmpdir, "export_combined.xlsx")
            
            # Export with combined sheet option (fmeda_per_fg=False)
            success = ExportService.export_to_excel(
                self.project,
                file_path,
                include_summary=True,
                include_custom=False,
                include_history=False,
                fmeda_per_fg=False
            )
            self.assertTrue(success)
            
            wb = openpyxl.load_workbook(file_path)
            self.assertIn("Overview", wb.sheetnames)
            self.assertNotIn("Power Supply", wb.sheetnames)
            self.assertNotIn("Change History", wb.sheetnames)
            self.assertIn("FMEDA Workspace", wb.sheetnames)
            
            ws_fmeda = wb["FMEDA Workspace"]
            # 1 extra column for "Functional Group" at the beginning -> 40 columns
            self.assertEqual(ws_fmeda.max_column, 40)
            self.assertEqual(ws_fmeda.cell(row=1, column=1).value, "Functional Group")
            self.assertEqual(ws_fmeda.cell(row=2, column=1).value, "Power Supply")
            self.assertEqual(ws_fmeda.cell(row=2, column=2).value, "R101")


if __name__ == "__main__":
    unittest.main()
