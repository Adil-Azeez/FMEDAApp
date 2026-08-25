import os
from pathlib import Path
from typing import Optional, List, Dict, Any
import json

from fmeda_tool.models import Project, Unit
from fmeda_tool.services.validation_service import ValidationService

# openpyxl import
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ExportService:
    """Service to handle Excel (.xlsx) and PDF report exports for FMEDA projects"""
    
    @staticmethod
    def export_to_excel(
        project: Project,
        filepath: str,
        include_summary: bool = True,
        include_custom: bool = True,
        include_history: bool = True,
        fmeda_per_fg: bool = True
    ) -> bool:
        """
        Exports the project FMEDA tables to a multi-sheet formatted Excel workbook.
        """
        if not OPENPYXL_AVAILABLE:
            # Fallback to CSV if openpyxl is not available
            return ExportService._export_to_csv_fallback(project, filepath)
            
        from fmeda_tool.services.calculation_service import CalculationService
        from fmeda_tool.models import FailureModeAssignment
        
        wb = openpyxl.Workbook()
        
        # Styles
        title_font = Font(name="Arial", size=16, bold=True, color="1F497D")
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        section_font = Font(name="Arial", size=12, bold=True, color="2E3B4E")
        label_font = Font(name="Arial", size=10, bold=True)
        value_font = Font(name="Arial", size=10)
        
        thin_side = Side(border_style="thin", color="D9D9D9")
        thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        # Sheet 1: Project Overview
        if include_summary:
            ws_overview = wb.active
            ws_overview.title = "Overview"
            ws_overview.views.sheetView[0].showGridLines = True
            
            # Title
            ws_overview["A1"] = "FMEDA Project Report"
            ws_overview["A1"].font = title_font
            
            # Metadata block
            ws_overview["A3"] = "Project Name:"
            ws_overview["B3"] = project.name
            ws_overview["A4"] = "Project Number:"
            ws_overview["B4"] = project.project_number or "N/A"
            ws_overview["A5"] = "Version:"
            ws_overview["B5"] = project.version
            ws_overview["A6"] = "Status:"
            ws_overview["B6"] = project.status.value.replace("_", " ").title() if project.status else "Draft"
            
            ws_overview["D3"] = "Safety Standard:"
            ws_overview["E3"] = project.safety_standard.value if project.safety_standard else "N/A"
            ws_overview["D4"] = "Target SIL:"
            ws_overview["E4"] = project.target_sil or "N/A"
            ws_overview["D5"] = "Reviewer:"
            ws_overview["E5"] = project.reviewer or "N/A"
            
            ws_overview["A7"] = "No Part Failure Def:"
            ws_overview["B7"] = (project.safety_context.no_part_failure_definition if project.safety_context else None) or "Not defined"
            ws_overview["B7"].alignment = Alignment(wrap_text=True)
            
            ws_overview["D7"] = "No Effect Failure Def:"
            ws_overview["E7"] = (project.safety_context.no_effect_failure_definition if project.safety_context else None) or "Not defined"
            ws_overview["E7"].alignment = Alignment(wrap_text=True)
            
            for r in range(3, 8):
                ws_overview[f"A{r}"].font = label_font
                ws_overview[f"B{r}"].font = value_font
                ws_overview[f"D{r}"].font = label_font
                ws_overview[f"E{r}"].font = value_font
                
            current_row = 9
            
            # Custom fields
            if include_custom and getattr(project, "custom_fields", None):
                ws_overview.cell(row=current_row, column=1, value="Custom Project Fields").font = section_font
                current_row += 2
                for k, v in project.custom_fields.items():
                    cell_k = ws_overview.cell(row=current_row, column=1, value=f"{k}:")
                    cell_k.font = label_font
                    cell_v = ws_overview.cell(row=current_row, column=2, value=v)
                    cell_v.font = value_font
                    current_row += 1
                current_row += 1
                
            # Summary metrics table
            ws_overview.cell(row=current_row, column=1, value="Calculated Safety Metrics Summary").font = section_font
            current_row += 2
            
            headers = ["Safety Parameter", "Value"]
            for c_idx, h in enumerate(headers, 1):
                cell = ws_overview.cell(row=current_row, column=c_idx, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Re-calculate to ensure latest values
            CalculationService.calculate_project(project)
            
            tot_dangerous = (project.dangerous_detected_rate or 0.0) + (project.dangerous_undetected_rate or 0.0)
            dc = (project.dangerous_detected_rate / tot_dangerous * 100.0) if tot_dangerous > 0.0 else 0.0
            
            metrics = [
                ("Total Failure Rate (FIT)", f"{project.total_failure_rate or 0.0:.4f}"),
                ("Safe Failure Fraction (SFF)", f"{project.sff or 0.0:.2f}%"),
                ("Dangerous Detected FIT", f"{project.dangerous_detected_rate or 0.0:.4f}"),
                ("Dangerous Undetected FIT", f"{project.dangerous_undetected_rate or 0.0:.4f}"),
                ("Safe Failure Rate FIT", f"{project.safe_failure_rate or 0.0:.4f}"),
                ("Average PFD (PFDavg)", f"{project.pfd_avg or 0.0:.6e}"),
                ("Diagnostic Coverage (DC)", f"{dc:.2f}%"),
                ("Achieved Safety Integrity Level", project.achieved_sil or "SIL 0")
            ]
            
            current_row += 1
            for param, val in metrics:
                cell_p = ws_overview.cell(row=current_row, column=1, value=param)
                cell_v = ws_overview.cell(row=current_row, column=2, value=val)
                cell_p.font = value_font
                cell_v.font = label_font
                cell_p.border = thin_border
                cell_v.border = thin_border
                cell_v.alignment = Alignment(horizontal="right" if "FIT" in param or "PFD" in param or "%" in param else "center")
                current_row += 1
                
            # Auto-adjust overview column widths
            for col in ws_overview.columns:
                max_len = 0
                for cell in col:
                    if cell.coordinate in ["B7", "E7"]:
                        continue
                    val_str = str(cell.value or '')
                    if '\n' in val_str:
                        lines = val_str.split('\n')
                        max_len = max(max_len, max(len(l) for l in lines))
                    else:
                        max_len = max(max_len, len(val_str))
                col_letter = get_column_letter(col[0].column)
                ws_overview.column_dimensions[col_letter].width = min(max(max_len + 3, 15), 45)
                
        # Sheets for Functional Group / FMEDA tables
        fmeda_headers = [
            "Component ID / Designator", "Status", "Function", "Value / Description",
            "Internal Part Number", "Fitted Status", "Component Type",
            "Failure Mode", "Failure-Mode %", "Base Failure Rate (FIT)",
            "Reliability Source", "Source Reference", "Environmental Profile",
            "Failure Effect / Deviation", "Diagnostic Function", "Failure Classification",
            "Diagnostic Measure ID", "Detection % (DC)", "DC Test Ref", "Mitigation",
            "Comments / Justification", "Review Status",
            "Proof Test A", "Proof Test B", "Proof Test C", "No Part / No Effect",
            "lambda (FIT)", "lambda_safe (FIT)", "lambda_dangerous (FIT)",
            "lambda_sd (FIT)", "lambda_su (FIT)", "lambda_dd (FIT)", "lambda_du (FIT)",
            "lambda_no_part (FIT)", "lambda_no_effect (FIT)", "SFF %", "DC %", "MTBF (h)", "MTTFd (y)"
        ]
        
        def write_fmeda_rows(ws, components, unit_name_opt=None):
            ws.views.sheetView[0].showGridLines = True
            
            headers_to_use = list(fmeda_headers)
            if unit_name_opt:
                # Add Functional Group column at beginning
                headers_to_use.insert(0, "Functional Group")
                
            for c_idx, h in enumerate(headers_to_use, 1):
                cell = ws.cell(row=1, column=c_idx, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.row_dimensions[1].height = 28
            
            row_idx = 2
            for comp in components:
                for fm_name, fm_percentage in comp.failure_modes.items():
                    assignment = next((a for a in comp.failure_mode_assignments if a.failure_mode_name == fm_name), None)
                    if not assignment:
                        assignment = FailureModeAssignment(
                            failure_mode_name=fm_name,
                            failure_rate_percentage=fm_percentage,
                            classification="not_evaluated",
                            dangerous_failure_percentage=100.0,
                            detection_percentage=0.0
                        )
                        
                    status_str = "Valid"
                    bg_color = "FFFFFF"  # White
                    
                    status, msgs = ValidationService.validate_row(assignment, comp)
                    if status == "error":
                        status_str = "Error"
                        bg_color = "F8D7DA"  # Light red
                    elif status == "warning":
                        status_str = "Warning"
                        bg_color = "FFF3CD"  # Light yellow
                        
                    row_fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
                    
                    local_fit = (comp.failure_rate or 0.0) * (fm_percentage / 100.0)
                    classif = getattr(assignment, "classification", "not_evaluated")
                    dp = assignment.dangerous_failure_percentage if assignment.dangerous_failure_percentage is not None else 100.0
                    det = assignment.detection_percentage if assignment.detection_percentage is not None else 0.0
                    
                    row_metrics = CalculationService.calculate_row_detailed(local_fit, classif, dp, det)
                    
                    # Resolve Deviation Name
                    dev_val = "N/A"
                    if assignment.deviation_id:
                        dev = next((d for d in project.deviations if d.id == assignment.deviation_id), None)
                        if dev:
                            dev_val = dev.name
                            
                    # Resolve Diagnostic Measure Name
                    dm_val = "N/A"
                    if assignment.diagnostic_measure_id:
                        dm = next((m for m in project.diagnostic_measures if m.id == assignment.diagnostic_measure_id), None)
                        if dm:
                            dm_val = dm.description
                            
                    # Resolve Mitigation Name
                    mit_val = "N/A"
                    if assignment.mitigation_id:
                        mit = next((m for m in project.mitigations if m.id == assignment.mitigation_id), None)
                        if mit:
                            mit_val = mit.name or mit.id
                            
                    class_map_rev = {
                        "not_evaluated": "Not Evaluated",
                        "safe_failure": "Safe Failure",
                        "dangerous_failure": "Dangerous Failure",
                        "no_effect_failure": "No Effect Failure",
                        "no_part_failure": "No Part Failure",
                        "diagnostic_function_failure": "Diagnostic Function Failure"
                    }
                    classif_val = class_map_rev.get(classif, "Not Evaluated")
                    
                    # Compile row values
                    row_values = [
                        comp.position,
                        status_str,
                        comp.function or "",
                        comp.value or "",
                        comp.internal_pn or "",
                        comp.fitted_status or "Fitted",
                        comp.type or "",
                        fm_name,
                        fm_percentage / 100.0,
                        comp.failure_rate or 0.0,
                        project.reliability_database_source or "MIL-HDBK-217F",
                        "Section 5",
                        project.environmental_profile or "Ground Benign (GB)",
                        dev_val,
                        assignment.diagnostic_function or "",
                        classif_val,
                        dm_val,
                        det / 100.0,
                        assignment.dc_test_ref or "",
                        mit_val,
                        assignment.notes or "",
                        (assignment.review_status or "Draft").title(),
                        getattr(assignment, "proof_test_a", 0.0) / 100.0,
                        getattr(assignment, "proof_test_b", 0.0) / 100.0,
                        getattr(assignment, "proof_test_c", 0.0) / 100.0,
                        "Yes" if getattr(assignment, "dont_care", False) else "No",
                        row_metrics["lambda"],
                        row_metrics["lambda_safe"],
                        row_metrics["lambda_dangerous"],
                        row_metrics["lambda_sd"],
                        row_metrics["lambda_su"],
                        row_metrics["lambda_dd"],
                        row_metrics["lambda_du"],
                        row_metrics["lambda_no_part"],
                        row_metrics["lambda_no_effect"],
                        row_metrics["sff"] / 100.0,
                        row_metrics["dc"] / 100.0,
                        row_metrics["mtbf"] if row_metrics["mtbf"] > 0 else "N/A",
                        row_metrics["mttfd"] if row_metrics["mttfd"] > 0 else "N/A"
                    ]
                    
                    if unit_name_opt:
                        comp_unit_name = ""
                        for u in project.units:
                            if any(c.id == comp.id for c in u.components):
                                comp_unit_name = u.name
                                break
                        row_values.insert(0, comp_unit_name)
                        
                    for col_idx, val in enumerate(row_values, 1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=val)
                        cell.font = value_font
                        cell.fill = row_fill
                        cell.border = thin_border
                        
                        # Apply number formatting
                        h_name = headers_to_use[col_idx - 1]
                        if "%" in h_name or "Distribution" in h_name or "Dangerous" in h_name or "Detection" in h_name or "Proof Test" in h_name or "SFF" in h_name or "DC" in h_name:
                            if isinstance(val, float):
                                cell.number_format = "0.0%"
                        elif "FIT" in h_name or "lambda" in h_name:
                            if isinstance(val, float):
                                cell.number_format = "0.0000"
                        elif "MTBF" in h_name:
                            if isinstance(val, float) and val > 0:
                                cell.number_format = "0.0e+0"
                        elif "MTTFd" in h_name:
                            if isinstance(val, float) and val > 0:
                                cell.number_format = "0.0"
                                
                        # Alignment
                        if h_name in ["Status", "Component ID / Designator", "Fitted Status", "Review Status", "No Part / No Effect", "Functional Group"]:
                            cell.alignment = Alignment(horizontal="center")
                        elif isinstance(val, (int, float)):
                            cell.alignment = Alignment(horizontal="right")
                            
                    row_idx += 1
                    
            # Auto-adjust column widths
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        if fmeda_per_fg:
            for unit in project.units:
                ws_title = unit.name[:30].replace("[", "").replace("]", "").replace("*", "").replace("?", "").replace("/", "").replace("\\", "")
                # Avoid title duplication if overview is named same
                if ws_title == "Overview" and include_summary:
                    ws_title = "Overview_FG"
                ws = wb.create_sheet(title=ws_title)
                write_fmeda_rows(ws, unit.components)
        else:
            # Combined sheet
            ws = wb.create_sheet(title="FMEDA Workspace")
            write_fmeda_rows(ws, [c for u in project.units for c in u.components], unit_name_opt="Functional Group")
            
        # Sheet: Change History
        if include_history and getattr(project, "change_history", None):
            ws_history = wb.create_sheet(title="Change History")
            ws_history.views.sheetView[0].showGridLines = True
            
            history_headers = ["Timestamp", "User", "Action", "Details"]
            for c_idx, h in enumerate(history_headers, 1):
                cell = ws_history.cell(row=1, column=c_idx, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            ws_history.row_dimensions[1].height = 24
            
            row_idx = 2
            for entry in project.change_history:
                ws_history.cell(row=row_idx, column=1, value=entry.get("timestamp", ""))
                ws_history.cell(row=row_idx, column=2, value=entry.get("user", ""))
                ws_history.cell(row=row_idx, column=3, value=entry.get("action", ""))
                ws_history.cell(row=row_idx, column=4, value=entry.get("details", ""))
                
                for col_c in range(1, 5):
                    cell_c = ws_history.cell(row=row_idx, column=col_c)
                    cell_c.font = value_font
                    cell_c.border = thin_border
                row_idx += 1
                
            # Auto-adjust column widths
            for col in ws_history.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws_history.column_dimensions[col_letter].width = max(max_len + 3, 15)
                
        # Remove default sheet if we added other sheets
        if len(wb.sheetnames) > 1 and "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
            
        wb.save(filepath)
        return True

    @staticmethod
    def _export_to_csv_fallback(project: Project, filepath: str) -> bool:
        """Fallback to CSV if openpyxl is not available"""
        import csv
        path = Path(filepath).with_suffix(".csv")
        with open(path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(["Project Name", project.name])
            writer.writerow(["Project Number", project.project_number])
            writer.writerow(["Version", project.version])
            writer.writerow(["SFF", f"{project.sff or 0.0:.2f}%"])
            writer.writerow([])
            
            for unit in project.units:
                writer.writerow([f"Functional Group: {unit.name}"])
                writer.writerow(["Designator", "Type", "Failure Mode", "Distribution %", "Local FIT"])
                for comp in unit.components:
                    for fm_name, fm_pct in comp.failure_modes.items():
                        local_fit = (comp.failure_rate or 0.0) * (fm_pct / 100.0)
                        writer.writerow([comp.position, comp.type, fm_name, f"{fm_pct}%", f"{local_fit:.4f}"])
                writer.writerow([])
        return True

    @staticmethod
    def export_to_pdf(project: Project, filepath: str) -> bool:
        """
        Exports the FMEDA report summary and signature approval block as a PDF file
        using PyQt6 QTextDocument print facility (requires zero external reportlab dependencies).
        """
        from PyQt6.QtGui import QTextDocument
        from PyQt6.QtPrintSupport import QPrinter
        from PyQt6.QtCore import QMarginsF
        import html
        
        sc = project.safety_context
        
        def esc(val, default="N/A", is_definition=False):
            if val is None or (isinstance(val, str) and not val.strip()):
                return "Not defined" if is_definition else default
            s = html.escape(str(val))
            return s.replace("\n", "<br/>")
            
        # HTML formatting for the document
        html_content = f"""
        <html>
        <head>
            <style>
                body {{ font-family: Arial, sans-serif; color: #333333; margin: 30px; }}
                h1 {{ color: #1F497D; font-size: 24px; border-bottom: 2px solid #1F497D; padding-bottom: 8px; }}
                h2 {{ color: #2E3B4E; font-size: 16px; margin-top: 25px; border-bottom: 1px solid #CCCCCC; padding-bottom: 4px; }}
                .meta-table {{ width: 100%; margin-bottom: 20px; border-collapse: collapse; }}
                .meta-table td {{ padding: 6px; font-size: 13px; }}
                .meta-label {{ font-weight: bold; color: #495057; width: 25%; }}
                .meta-value {{ color: #212529; }}
                .summary-table {{ width: 100%; border-collapse: collapse; margin-top: 10px; }}
                .summary-table th {{ background-color: #1F497D; color: white; padding: 8px; font-size: 13px; text-align: left; }}
                .summary-table td {{ padding: 8px; border: 1px solid #D9D9D9; font-size: 13px; }}
                .summary-value {{ font-weight: bold; text-align: right; }}
                .signature-block {{ margin-top: 50px; width: 100%; border-collapse: collapse; }}
                .signature-block td {{ padding: 15px; font-size: 13px; vertical-align: bottom; }}
                .sig-line {{ border-bottom: 1px solid #333333; width: 250px; height: 40px; }}
                .fg-table {{ width: 100%; border-collapse: collapse; margin-top: 10px; }}
                .fg-table th {{ background-color: #2E3B4E; color: white; padding: 6px; font-size: 11px; }}
                .fg-table td {{ padding: 6px; border: 1px solid #E5E5E5; font-size: 11px; }}
            </style>
        </head>
        <body>
            <h1>FMEDA Engineering Analysis Report</h1>
            
            <table class="meta-table">
                <tr>
                    <td class="meta-label">Project Name:</td>
                    <td class="meta-value">{project.name}</td>
                    <td class="meta-label">Safety Standard:</td>
                    <td class="meta-value">{project.safety_standard.value if project.safety_standard else 'N/A'}</td>
                </tr>
                <tr>
                    <td class="meta-label">Project Number:</td>
                    <td class="meta-value">{project.project_number or 'N/A'}</td>
                    <td class="meta-label">Target SIL:</td>
                    <td class="meta-value">{project.target_sil or 'N/A'}</td>
                </tr>
                <tr>
                    <td class="meta-label">Version:</td>
                    <td class="meta-value">{project.version}</td>
                    <td class="meta-label">Reviewer:</td>
                    <td class="meta-value">{project.reviewer or 'N/A'}</td>
                </tr>
                <tr>
                    <td class="meta-label">Status:</td>
                    <td class="meta-value">{project.status.value.replace("_", " ").title() if project.status else 'Draft'}</td>
                    <td class="meta-label">Date Generated:</td>
                    <td class="meta-value">{project.updated_at.strftime('%Y-%m-%d %H:%M:%S')}</td>
                </tr>
            </table>
            
            <h2>Safety Context & Parameters</h2>
            <table class="meta-table">
                <tr>
                    <td class="meta-label" style="width: 35%;">Safety Function Name:</td>
                    <td class="meta-value">{esc(sc.safety_function_name if sc else None)}</td>
                </tr>
                <tr>
                    <td class="meta-label" style="width: 35%;">Safety Function Description:</td>
                    <td class="meta-value">{esc(sc.safety_function_description if sc else None)}</td>
                </tr>
                <tr>
                    <td class="meta-label" style="width: 35%;">Safe State Definition:</td>
                    <td class="meta-value">{esc(sc.safe_state if sc else None)}</td>
                </tr>
                <tr>
                    <td class="meta-label" style="width: 35%;">Dangerous State Definition:</td>
                    <td class="meta-value">{esc(sc.dangerous_state if sc else None)}</td>
                </tr>
                <tr>
                    <td class="meta-label" style="width: 35%;">No Part Failure Definition:</td>
                    <td class="meta-value">{esc(sc.no_part_failure_definition if sc else None, is_definition=True)}</td>
                </tr>
                <tr>
                    <td class="meta-label" style="width: 35%;">No Effect Failure Definition:</td>
                    <td class="meta-value">{esc(sc.no_effect_failure_definition if sc else None, is_definition=True)}</td>
                </tr>
                <tr>
                    <td class="meta-label" style="width: 35%;">Safety Architecture:</td>
                    <td class="meta-value">{esc(sc.safety_architecture if sc else None)}</td>
                </tr>
                <tr>
                    <td class="meta-label" style="width: 35%;">Operating Mode:</td>
                    <td class="meta-value">{esc(sc.operating_mode if sc else None)}</td>
                </tr>
                <tr>
                    <td class="meta-label" style="width: 35%;">Safety Boundary Definition:</td>
                    <td class="meta-value">{esc(sc.safety_boundary if sc else None)}</td>
                </tr>
                <tr>
                    <td class="meta-label" style="width: 35%;">External Sensor Included:</td>
                    <td class="meta-value">{esc("Yes" if (sc and sc.external_sensor_included) else "No")}</td>
                </tr>
                <tr>
                    <td class="meta-label" style="width: 35%;">Reliability DB Source:</td>
                    <td class="meta-value">{esc(project.reliability_database_source)}</td>
                </tr>
                <tr>
                    <td class="meta-label" style="width: 35%;">Environmental Profile:</td>
                    <td class="meta-value">{esc(project.environmental_profile)}</td>
                </tr>
            </table>
            
            <h2>Calculated Safety Metrics</h2>
            <table class="summary-table">
                <tr>
                    <th>Metric Parameter</th>
                    <th>Headings</th>
                </tr>
                <tr>
                    <td>Total Failure Rate (FIT)</td>
                    <td class="summary-value">{project.total_failure_rate or 0.0:.4f} FIT</td>
                </tr>
                <tr>
                    <td>Safe Failure Fraction (SFF)</td>
                    <td class="summary-value">{project.sff or 0.0:.2f}%</td>
                </tr>
                <tr>
                    <td>Dangerous Detected FIT</td>
                    <td class="summary-value">{project.dangerous_detected_rate or 0.0:.4f} FIT</td>
                </tr>
                <tr>
                    <td>Dangerous Undetected FIT</td>
                    <td class="summary-value">{project.dangerous_undetected_rate or 0.0:.4f} FIT</td>
                </tr>
                <tr>
                    <td>Safe Failure Rate FIT</td>
                    <td class="summary-value">{project.safe_failure_rate or 0.0:.4f} FIT</td>
                </tr>
                <tr>
                    <td>Average Probability of Failure on Demand (PFDavg)</td>
                    <td class="summary-value">{project.pfd_avg or 0.0:.6e}</td>
                </tr>
                <tr>
                    <td>Achieved Safety Integrity Level</td>
                    <td class="summary-value" style="color: #198754;">{project.achieved_sil or 'SIL 0'}</td>
                </tr>
            </table>
            
            <h2>Functional Groups Summary</h2>
            <table class="summary-table">
                <tr>
                    <th>Group Name</th>
                    <th>Failure Rate (FIT)</th>
                    <th>SFF %</th>
                    <th>DC %</th>
                </tr>
        """
        
        for u in project.units:
            html_content += f"""
                <tr>
                    <td>{u.name}</td>
                    <td>{u.total_failure_rate or 0.0:.4f}</td>
                    <td>{u.safe_failure_fraction or 0.0:.2f}%</td>
                    <td>{u.diagnostic_coverage or 0.0:.2f}%</td>
                </tr>
            """
            
        html_content += """
            </table>
            
            <div style="page-break-before: always;"></div>
            <h1>Approvals & Verification Signatures</h1>
            <p>By signing below, the reviewer and creator verify that the failure mode distribution rates, diagnostic coverage validations, and calculated Safety Integrity Level metrics meet standard criteria.</p>
            
            <table class="signature-block">
                <tr>
                    <td>
                        <div class="sig-line"></div>
                        <div style="font-weight: bold; margin-top: 5px;">Created By:</div>
                        <div style="color: #666666;">Safety Engineer</div>
                    </td>
                    <td>
                        <div class="sig-line"></div>
                        <div style="font-weight: bold; margin-top: 5px;">Verified / Reviewed By:</div>
                        <div style="color: #666666;">Independent Safety Assessor</div>
                    </td>
                </tr>
            </table>
        </body>
        </html>
        """
        
        doc = QTextDocument()
        doc.setHtml(html_content)
        
        printer = QPrinter(QPrinter.PrinterMode.HighResolution)
        printer.setOutputFormat(QPrinter.OutputFormat.PdfFormat)
        printer.setOutputFileName(filepath)
        
        layout = printer.pageLayout()
        layout.setMargins(QMarginsF(15, 15, 15, 15))
        printer.setPageLayout(layout)
        
        if hasattr(doc, "print"):
            doc.print(printer)
        else:
            doc.print_(printer)
        return True
